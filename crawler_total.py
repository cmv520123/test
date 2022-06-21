# keyword = '保力達'
def crawler_all(keyword):
#     from bs4 import BeautifulSoup
#     from openpyxl import Workbook
#     import parsel #數據解析模組
#     import requests #數據請求模組
#     import csv
#     import json
#     from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
#     import re
#     import pandas as pd
#     from json.decoder import JSONDecodeError
#     import os
    # keywords_list = ['老協珍','田原香','農純鄉','芳茲','娘家','純煉']
    # 為防爬蟲監控 所以改成自己的user-agent
    # headers = {
    #     '改成自己的user-agent'
    # }
    # #為防爬蟲監控 所以改成自己的user-agent
    # raw_folder = r'C:\Users\User\Desktop\0614_datasets\ '+keyword
    # txt_folder = r'C:\Users\User\Desktop\0614_datasets\ '

    os.makedirs(keyword+'_data',exist_ok = True)
    os.makedirs(keyword+'_analysis',exist_ok = True)
    headers = {
        'user-agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.0.0 Safari/537.36'
    }

    ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

    wb = Workbook()
    ws = wb.active

    title =["文章標題","發佈日期",'關鍵字',"文章內容"]
    ws.append(title)
    web = '媽咪拜'

    article_url=[]
    article_url_2 = []

    for i in range(1,20):
        try:
            url = 'https://mamibuy.com.tw/search/'+keyword+'?p='+str(i)+'&s=1'
            response = requests.get(url=url , headers = headers)
            selector = parsel.Selector(response.text) # 把獲取下來的html字符串數據，轉成selector可解析的對象
        # print (selector) #<Selector xpath=None data='<html lang="zh-CN" class="ua-mac ua-w...'>
        # css選擇器：就是根據標籤屬性內容，提取相關數據
            lis = selector.css('.well .link-black') #第一次提取，獲取所有li標籤 返回列表
        # print(lis)
            for li in lis :
                href = li.css('a::attr(href)').get() #詳情頁 獲取attr屬性
                article_url.append(href)
        except KeyError as k :
            print(k)     
        except AttributeError as attr:
            print(attr)
        except NameError as name:
            print(name)
        except IndexError as index:
            print(index)
        except JSONDecodeError :
            print(JSONDecodeError)
        # print(article_url)
        # break
    # print(len(article_url))

    for i in article_url :
        y = i.replace('https://mamibuy.com.tw','')
        z = 'https://mamibuy.com.tw'+ y
        article_url_2.append(z)
        # print(z)

    print(web + '總共有'+str(len(article_url_2))+'個網址') 

    # 爬取所有文章的內文
    num = 0 
    for url in article_url_2:
        # try:
        # print(url)
        # 1.發送請求

        res = requests.get(url=url,headers=headers)

        soup = BeautifulSoup(res.text,"lxml")
        json_file = soup.select("script[type='application/ld+json']")
        # j = json_file[1]
        # print(json_file) 
        # print(type(json_file))
        #<class 'bs4.element.Tag'>
        try:
            for i in json_file:

                targets = json.loads(i.get_text(strip=True),strict=False)
                # strict=False 不要那麼嚴謹 字典取不到的 就算了
                # print(targets['articleBody'])
                titles_list=["headline","datePublished",'keywords',"articleBody"]
                # title =["文章標題","發佈日期",'關鍵字',"文章內容"]
                article_info = []
                try :
                    for title in titles_list:                
                        y = targets[title]
                        y= ILLEGAL_CHARACTERS_RE.sub(r'', y)
                        article_info.append(y)
                        # print(title,':',y)    
                except KeyError as k :
                    print(k)             
                except AttributeError as attr:
                    print(attr)
                except NameError as name:
                    print(name)
                except IndexError as index:
                    print(index)
                except JSONDecodeError :
                    print(JSONDecodeError)
                if len(article_info) != 0 :
                    num+=1
                    ws.append(article_info)
                    print(web + '已完成寫入' + str(num) + '篇文章')
        except KeyError as k :
            print(k)       
        except AttributeError as attr:
            print(attr)
        except NameError as name:
            print(name)
        except IndexError as index:
            print(index)
        except JSONDecodeError :
            print(JSONDecodeError)

    wb.save(keyword +'_'+ web + '_網路爬蟲.xlsx')

    import pandas as pd
    df = pd.read_excel(keyword +'_'+ web + '_網路爬蟲.xlsx') #請自行依檔案位置調整
    print(keyword+'_'+web+ '_總共有'+str(len(df))+'篇文章')

    #媽咪愛爬蟲

    from bs4 import BeautifulSoup
    from openpyxl import Workbook
    import parsel #數據解析模組
    import requests #數據請求模組
    import csv
    import json
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
    import openpyxl
    import re

    wb = Workbook()
    ws = wb.active

    title = ["文章ID","文章標題","文章內容","留言數","留言內容"]

    ws.append(title)
    ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
    web = '媽咪愛'

    url = 'https://mamilove.com.tw/v3/questions/search?query='+str(keyword)+'&per_page=10000'

    res = requests.get(url)
    res_json = res.json()
    love_json = res_json['data']
        
    num =0 
    for article in love_json:
        try:
            article_info = []
            titles_list=["question_id","subject","content"]


            for title in titles_list:                
                y = article[title]

                article_info.append(y)
                # print(title,':',y)
                
            d = article["question_meta"]["answer_count"]
            title_2 = 'answer_count'
            article_info.append(d)
            # print(title_2,':',d)


            #     for j in love_json:
            #         article_ID = love_json["question_id"]
            article_ID = article['question_id']
            url = 'https://mamilove.com.tw/v3/question/'+str(article_ID)+'/answers?per_page=200'
            message = requests.get(url)
            message_json = message.json()
            clean_message = message_json['data']
            pure_message = ''
            for w in clean_message:
                pure_message += w['content']
                pure_message = str(pure_message)
                pure_message= ILLEGAL_CHARACTERS_RE.sub(r'', pure_message)
            #             pure_message = pure_message.replace('[\000-\010]','').replace('|[\013-\014]','').replace('|[\016-\037]','')
            article_info.append(pure_message)
            title_3 = 'message'
            # print(title_3,':',pure_message)
            num += 1
            ws.append(article_info)
            print(web + '已完成寫入' + str(num) + '篇文章')
        except KeyError as k :
            print(k)        
        except AttributeError as attr:
            print(attr)
        except NameError as name:
            print(name)
        except IndexError as index:
            print(index)
        except JSONDecodeError :
            print(JSONDecodeError) 

    #         print(article_info)
    #         break

    wb.save(keyword +'_'+ web + '_網路爬蟲.xlsx')

    df = pd.read_excel(keyword + '_'+ web +'_網路爬蟲.xlsx') #請自行依檔案位置調整
    print(keyword+'_'+web+ '_總共有'+str(len(df))+'篇文章')

    #爬取Dcard深卡貼文標題及內文
    import requests as req
    from openpyxl import Workbook
    import numpy as np
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
    import re
    import pandas as pd

    #新增Excel及工作表
    wb = Workbook()
    ws = wb.active
    ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
    #新增title
    title = ['發佈日期','文章標題','文章內容','留言數','喜歡數','標籤']
    ws.append(title)
    web = 'Dcard深卡'

    #爬取的日期 2019/1/1~2022/12/28
    years = ['2019','2020','2021','2022']
    # dates = ['01','15','28']
    dates = ['01','28']

    num = 0
    for year in years:    
        for month in range(1,13,3):
            for date in dates:
                try:

                    # print(year+'-'+str(month)+'-'+date)
                    #不斷更改網址 不斷前往網站爬取
                    url = 'https://tw.observer/api/posts?hot=0&before='+year+'-'+str(month)+'-'+date+'&term='+keyword
                    #https://tw.observer/api/posts?hot=0&before=2022-05-09&term=%滴雞精
                    r = req.get(url,headers=headers)
                    #以json格式讀取
                    rj = r.json()
                    #拆解資料格式-字典
                    #{'data': {'posts': 
                    # [{'id': 230313090, 'title': '目前懷孕16週的媽媽', 
                    # 'content': '懷孕至目前16週了\n懷孕之前是166/59，結果前三個月非常嚴重的孕吐讓體重直直落瘦到51還伴隨著些許落紅\n好不容易熬過但食慾還是不見起色\n目前食量大約5顆水餃一餐，就算沒吃也不覺得餓，覺得對不起寶寶但逼自己多吃一點又開始吐\n\n\n而且還不幸感冒😷頭痛欲裂\n有去婦產科看醫生快一個月還好不了\n也不敢一直求醫怕吃太多藥會影響寶寶\n問醫生怎麼辦，他只說那妳要趕快好啊⋯⋯\n\n\n因為食慾一直不振\n不知道有沒有媽媽，可以給我一些建議\n我是不是該開始喝滴雞精或其他食物呢？', 
                    # 'excerpt': '懷孕至目前16週了，懷孕之前是166/59，結果前三個月非常嚴重的孕吐讓體重直直落瘦到51還伴隨著些許落紅，好不容易熬過但食慾還是不見起色，目前食量大約5顆水餃一餐，就算沒吃也不覺得餓，覺得對不起寶寶但逼自己多吃一點又開始吐', 
                    # 'createdAt': '2018-12-20T10:10:14.846000Z', 
                    # 'updatedAt': '2018-12-20T10:10:14.846000', 
                    # 'commentCount': 19, 'likeCount': 23, 'forumName': '結婚', 
                    # 'forumAlias': 'marriage', 'gender': 'F', 'school': '朝陽科技大學', 
                    # 'hidden': False, 'media': [], 'score': '2018-12-20T15:53:46.484000Z', 'order': '2018-12-20T10:10:14.846000Z'},
                    rjs = rj['data']['posts']
                    rx = rjs
                    #將所需資料爬取後 存進posts
                    for i in range(len(rx)):
                        posts = []
                        posts.append(rx[i]['order'])
                        posts.append(rx[i]['title'])
                        content = rx[i]['content']
                        content = ILLEGAL_CHARACTERS_RE.sub(r'', content)
                        posts.append(content)
                        posts.append(rx[i]['commentCount'])
                        posts.append(rx[i]['likeCount'])
                        posts.append(rx[i]['forumName'])
                        # title = ['發佈時間','文章標題','內文','留言數','喜歡數','標籤']
                        #將資料存進Excel
                        num +=1
                        ws.append(posts)
                        print(web + '已完成寫入'+ str(num) + '篇文章')
                    # print(posts)
                except KeyError as k :
                    print(k)
                except AttributeError as attr:
                    print(attr)
                except NameError as name:
                    print(name)
                except IndexError as index:
                    print(index)
                except JSONDecodeError :
                    print(JSONDecodeError)

    #要加save才能輸出成Excel
    wb.save(keyword+ '_yet_' + web + '_網路爬蟲.xlsx')

    #讀取剛爬取好的檔案
    df = pd.read_excel(keyword+ '_yet_' + web + '_網路爬蟲.xlsx') #請自行依檔案位置調整

    #查看是否有重複值
    cd1 = df.duplicated('文章標題').sum()
    print('重複文章有'+str(cd1)+'篇')

    #以文章標題作為標的 刪除重複值
    drop_data = df.drop_duplicates(subset=['文章標題'])
    # print('總共有'+str(len(drop_data))+'篇文章')
    #要加to_excel才能輸出成Excel
    drop_data.to_excel(keyword+'_ok_'+'_'+ web +'_網路爬蟲.xlsx',index=False)

    df = pd.read_excel(keyword+ '_ok' +'_'+ web + '_網路爬蟲.xlsx') #請自行依檔案位置調整
    print(keyword+'_'+web+ '_總共有'+str(len(df))+'篇文章')

    # PTT網路爬蟲
    # 使用方式可更改 1.關鍵字 2.可爬多個看板(自行選擇) 3.爬取的頁數 
    # p.s. 推 平 噓 留言皆已分開 / 資料最終以excel形式存放
    from bs4 import BeautifulSoup
    import requests
    import json
    import csv
    from openpyxl import Workbook
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
    import re

    ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

    wb = Workbook()
    ws = wb.active

    title =["作者",'看板',"文章標題","發佈日期","文章內容","推推留言",'平平留言','噓噓留言']

    ws.append(title)

    web = 'PTT'
    # board_names = ['Gossiping', 'Stock', 'Baseball', 'C_Chat', 'LoL', 'NBA', 'Lifeismoney', 'Elephants', 'HatePolitics', 'Military', 'car', 'Tech_Job', 'Beauty', 'home-sale', 'KoreaStar', 'Lions', 'BabyMother', 'DIABLO', 'MobileComm', 'Boy-Girl', 'movie', 'sex', 'Monkeys', 'WomenTalk', 'PC_Shopping', 'marriage', 'Badminton', 'KoreaDrama', 'AllTogether', 'DigiCurrency', 'Kaohsiung', 'Tainan', 'Guardians', 'PlayStation', 'basketballTW', 'TaichungBun', 'japanavgirls', 'KR_Entertain', 'creditcard', 'joke', 'NSwitch', 'Steam', 'TaiwanDrama', 'China-Drama', 'CVS', 'CFantasy', 'Marginalman', 'marvel', 'e-shopping', 'Japandrama', 'SportLottery', 'Insurance', 'iOS', 'EAseries', 'HardwareSale', 'forsale', 'BeautySalon', 'Hsinchu', 'AC_In', 'KoreanPop', 'Gamesale', 'Soft_Job', 'StupidClown', 'biker', 'MacShop', 'watch', 'TW_Entertain', 'PuzzleDragon', 'YuanChuang', 'Headphone', 'Salary', 'studyteacher', 'Option', 'FORMULA1', 'PublicServan', 'CarShop', 'Tennis', 'ToS', 'MakeUp', 'FATE_GO', 'fastfood', 'PokemonGO', 'MuscleBeach', 'BaseballXXXX', 'Brand', 'DMM_GAMES', 'NBA_Film', 'ONE_PIECE', 'E-appliance', 'MLB', 'Taoyuan', 'Bank_Service', 'nb-shopping', 'WOW', 'nCoV2019', 'YUGIOH', 'cookclub', 'Hearthstone', 'SakaTalk', 'Gov_owned', 'miHoYo', 'mobilesales', 'cat', 'Wanted', 'BabyProducts', 'gay', 'give', 'Teacher', 'Examination', 'GetMarry', 'SuperBike', 'Finance', 'hypermall', 'Palmar_Drama', 'Food', 'job', 'Arknights', 'Key_Mou_Pad', 'Digitalhome', 'UmaMusume', 'MH', 'Vtuber', 'FTV', 'Railway', 'Hip-Hop', 'HelpBuy', 'feminine_sex', 'Coffee']
    board_names = ['Gossiping', 'Beauty','BabyMother','Boy-Girl',
        'WomenTalk','sex','marriage','AllTogether','KoreaDrama',
        'TaichungBun','joke','Marginalman','e-shopping','BeautySalon'
        'AC_In','Gamesale','MakeUp','fastfood','nCoV2019','BabyProducts',
        'cookclub','Food','Key_Mou_Pad']
    # # 基本參數
    # # url = "https://www.ptt.cc/bbs/"+board_names+"/search?q="+ keyword
    # # for j in range(2):
    payload = {
        'from': '/bbs/Gossiping/index.html',
        'yes': 'yes'
    }
    data = []   # 全部文章的資料
    num = 0 #

    # # 用session紀錄此次使用的cookie，就是按過一次18，第二次就不必再按
    rs = requests.session()
    response = rs.post("https://www.ptt.cc/ask/over18", data=payload)
    page_urls = []
    num = 0
    for board_name in board_names:
        #range 可以調整爬蟲的頁數
        for i in range(1,11):
            url = "https://www.ptt.cc/bbs/"+board_name+'/search?page='+str(i)+'&q='+keyword
        # 爬取兩頁
                # get取得頁面的HTML
                # print(url)
            response = rs.get(url)
            soup = BeautifulSoup(response.text, "html.parser")
            # print(soup.prettify()) 

            # 找出每篇文章的連結
            links = soup.find_all("div", class_="title")
            for link in links:
                # 如果文章已被刪除，連結為None
                if link.a != None:
                    article_data = {}   # 單篇文章的資料
                    page_url = "https://www.ptt.cc/"+link.a["href"]
                    if page_url not in page_urls :
                        page_urls.append(page_url)

    # print(page_urls)
    print(web + '總共有' +str(len(page_urls))+'個網址')
    # page_urls =['https://www.ptt.cc//bbs/Gossiping/M.1648429698.A.8CE.html', 'https://www.ptt.cc//bbs/Gossiping/M.1644230304.A.FD3.html', 'https://www.ptt.cc//bbs/Gossiping/M.1643785151.A.374.html', 'https://www.ptt.cc//bbs/Gossiping/M.1642565644.A.5FF.html', 'https://www.ptt.cc//bbs/Gossiping/M.1640247990.A.53C.html', 'https://www.ptt.cc//bbs/Gossiping/M.1639223413.A.BA6.html', 'https://www.ptt.cc//bbs/Gossiping/M.1639222570.A.5DC.html', 'https://www.ptt.cc//bbs/Gossiping/M.1639222163.A.CC0.html', 'https://www.ptt.cc//bbs/Gossiping/M.1638419405.A.A28.html', 'https://www.ptt.cc//bbs/Gossiping/M.1635037526.A.519.html', 'https://www.ptt.cc//bbs/Gossiping/M.1633216902.A.ED8.html', 'https://www.ptt.cc//bbs/Gossiping/M.1631795899.A.ABC.html', 'https://www.ptt.cc//bbs/Gossiping/M.1631780466.A.21B.html', 'https://www.ptt.cc//bbs/Gossiping/M.1630490246.A.AE5.html', 'https://www.ptt.cc//bbs/Gossiping/M.1630073547.A.292.html', 'https://www.ptt.cc//bbs/Gossiping/M.1629223525.A.12F.html', 'https://www.ptt.cc//bbs/Gossiping/M.1626313436.A.AE0.html', 'https://www.ptt.cc//bbs/Gossiping/M.1624774427.A.257.html', 'https://www.ptt.cc//bbs/Gossiping/M.1623123641.A.3AD.html', 'https://www.ptt.cc//bbs/Gossiping/M.1622704613.A.BD5.html', 'https://www.ptt.cc//bbs/Gossiping/M.1621039775.A.F12.html', 'https://www.ptt.cc//bbs/Gossiping/M.1618761431.A.F32.html', 'https://www.ptt.cc//bbs/Gossiping/M.1618618396.A.AAF.html']

    try :
        article_data = {}   # 單篇文章的資料

        # 進入文章頁面
        for n in page_urls :
            response = rs.get(n)
            result = BeautifulSoup(response.text, "html.parser")
            # print(result)
            # break
            # print(soup.prettify()) #確認單個頁面標題那些

            # 找出作者、標題、時間、留言
            main_content = result.find("div", id="main-content")
            article_info = main_content.find_all("span", class_="article-meta-value")
            author = article_info[0].string  # 作者
            board = article_info[1].string #看板
            title = article_info[2].string  # 標題
            time = article_info[3].string   # 時間
            # if len(article_info) != 0: #不等於就跳出(跑一次就跳完)
            #     author = article_info[0].string  # 作者
            #     title = article_info[2].string  # 標題
            #     time = article_info[3].string   # 時間
            # else:
            #     author = "無"  # 作者
            #     title = "無"  # 標題
            #     time = "無"   # 時間

            # print(author)
            # print(title)
            # print(time)
            # break
                # break

            article_data["author"] = author
            article_data['board'] = board
            article_data["title"] = title
            article_data["time"] = time #到時候要取出容器資料-作者文章時間
            # print(article_data)
            # 將整段文字內容抓出來，因為沒有div容器包著
            all_text = main_content.text
            # 以--切割，抓最後一個--前的所有內容
            pre_texts = all_text.split("--")[:-1]
            # 將前面的所有內容合併成一個
            one_text = "--".join(pre_texts)
            # 以\n切割，第一行標題不要
            texts = one_text.split("\n")[1:]
            # 將每一行合併
            content = "\n".join(texts)
            # print(content)

            article_data["content"] = content  #到時候要取出容器資料-內容
            #     #留言跟換頁的code
            #     # 一種留言一個列表，這邊是來顯示、推、平、噓
            comment_dic = {}
            push_dic = []
            arrow_dic = []
            shu_dic = []

            # 抓出所有留言
            comments = main_content.find_all("div", class_="push")
            #BeautifulSoup不能使用.text取文字，需要使用參數轉換成字串
            for comment in comments:
                push_tag = comment.find(
                    "span", class_="push-tag").string   # 分類標籤
                push_userid = comment.find(
                    "span", class_="push-userid").string  # 使用者ID
                push_content = comment.find(
                    "span", class_="push-content").string   # 留言內容
                push_time = comment.find(
                    "span", class_="push-ipdatetime").string   # 留言時間

                # print(push_tag, push_userid, push_content, push_time)

                # dict1 = {'push_tag':push_tag,"push_userid": push_userid,
                #         "push_content" : push_content, "push_time": push_time}

                dict1 = {"push_userid": push_userid,
                        "push_content" : push_content, "push_time": push_time}
                if push_tag == "推 " :
                    push_dic.append(dict1)
                elif push_tag == "→ ":
                    arrow_dic.append(dict1)
                else:
                    shu_dic.append(dict1)
                # else:
                #     break
            # print(dict1)
            #         #辨識字串內容，轉成文字在附加到字典裡面
                # print(dict1["push_content"])
            # print(type(dict1))

            # # # # #這邊輸出會有一堆代碼，然後還要還需要輸出給那3個表

            # print(push_dic)
            # print(arrow_dic)
            # print(shu_dic)
            # print("--------")

            comment_dic["推"] = push_dic
            comment_dic["→"] = arrow_dic
            comment_dic["噓"] = shu_dic
            article_data["comment"] = comment_dic
            #comment_dict 存放著已分好的推 平 噓 三種評論內容
            # print(comment_dic)

            # print(article_data) #看一下未整理的，但存在這個容器內的資料
            #{'author': 'id520 (也無風雨也無晴)', 
            # 'title': '[問卦] 超商有沒有甚麼益生菌或優格產品可以推', 
            # 'time': 'Mon Mar 28 09:08:15 2022', 
            # 'content': '\n  剛好有超商一百元抵用券 想說來顧個健康\n\n  有沒有甚麼益生菌或是優格產品值得買的\n\n  今天台股又大跌了\n\n  上禮拜五賠錢賣興富發買進台積電\n\n  結果今天又套牢了  我是不是真的被主力盯上了\n\n\n', 
            # 'comment': {'推': [{'push_tag': '推 ', 'push_userid': 'lpbrother', 'push_content': ': 養樂多', 'push_time': '101.136.131.159 03/28 09:14\n'}, {'push_tag': '推 ', 'push_userid': 'likebadday', 'push_content': ': 福樂黃色包裝小鮮奶優酪，口感扎實綿', 'push_time': '  1.160.196.224 03/30 02:47\n'}], '→': [{'push_tag': '→ ', 'push_userid': 'ansfan', 'push_content': ': 葡眾', 'push_time': '  49.216.130.81 03/28 09:08\n'}, {'push_tag': '→ ', 'push_userid': 'likebadday', 'push_content': ': 密、腸胃蠕動順', 'push_time': '  1.160.196.224 03/30 02:47\n'}], '噓': []}}

            data.append(article_data)#將所有文章內容塞進"data"裡面

            for m in data:
                author = m['author'] 
                board = m['board']
                title = m['title']
                time = m['time']
                content =m['content']
                num += 1
                comment_push = []
                for a in range(len(m['comment']['推'])):
                    push = str(m['comment']['推'][a]['push_content'])
                    comment_push.append(push)
                comment_push=str(comment_push)

                comment_flat = []
                for b in range(len(m['comment']['→'])):
                    flat = str(m['comment']['→'][b]['push_content'])
                    comment_flat.append(flat)
                comment_flat=str(comment_flat)

                comment_shu = []
                for c in range(len(m['comment']['噓'])):
                    shu = str(m['comment']['噓'][c]['push_content'])
                    comment_shu.append(shu)
                comment_shu=str(comment_shu)

                items = [author,board,title,time,content,comment_push,comment_flat,comment_shu]
                
                excel_contents = []
                for w in items:
                    w= ILLEGAL_CHARACTERS_RE.sub(r'', w)
                    excel_contents.append(w)
                
                # print(excel_contents)
                print(web + '已完成寫入' + str(num) + '篇文章')
                ws.append(excel_contents)
                break
                # print(excel_contents)
                # print(excel_contents)
                # print(author1,title,time,content,comment_push,comment_flat,comment_shu,sep='===')
    except KeyError:
        print(KeyError)          
    except AttributeError as attr:
        print(attr)
    except NameError as name:
        print(name)
    except IndexError as index:
        print(index)
    except JSONDecodeError :
        print(JSONDecodeError)
    wb.save(keyword + '_'+ web + '_網路爬蟲.xlsx')

    df = pd.read_excel(keyword + '_'+ web +'_網路爬蟲.xlsx') #請自行依檔案位置調整
    print(keyword+'_'+web+ '_總共有'+str(len(df))+'篇文章')


    from bs4 import BeautifulSoup
    from openpyxl import Workbook
    import parsel #數據解析模組
    import requests #數據請求模組
    import csv
    import json
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
    import re

    ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

    wb = Workbook()
    ws = wb.active

    title =["文章標題","文章分類","發佈日期","文章內容","文章網址"]

    ws.append(title)

    web = 'TVBS'

    #為防爬蟲監控 所以改成自己的user-agent

    # header = {
    #     '改成自己的網址'
    # }
    article_url=[]

    for i in range(1,11):
        url = 'https://news.tvbs.com.tw/news/searchresult/'+keyword+'/news/'+str(i)
        response = requests.get(url=url , headers = headers)
        selector = parsel.Selector(response.text) # 把獲取下來的html字符串數據，轉成selector可解析的對象
        # print (selector) #<Selector xpath=None data='<html lang="zh-CN" class="ua-mac ua-w...'>
        # css選擇器：就是根據標籤屬性內容，提取相關數據
        lis = selector.css('.news_list .list ul li') #第一次提取，獲取所有li標籤 返回列表
        # print(lis)
        for li in lis :
            href = li.css('a::attr(href)').get() #詳情頁 獲取attr屬性
            article_url.append(href)

    print(web +'總共有'+ str(len(article_url)) +'個網址') 
    # #爬取所有文章的網址


    #爬取所有文章的內文
    num = 0
    for url in article_url:
        try:
        # print(url)
        # 1.發送請求

            res = requests.get(url=url,headers=headers)

            soup = BeautifulSoup(res.text,"lxml")
            json_file = soup.select("script[type='application/ld+json']")
            # print(type(json_file))
            #<class 'bs4.element.ResultSet'>
            for i in json_file:
                targets = json.loads(i.get_text(strip=True),strict=False)
                #strict=False 不要那麼嚴謹 字典取不到的 就算了
                titles_list=["headline","articleSection","datePublished","articleBody","mainEntityOfPage"]
                #title =["文章標題","文章分類","發佈日期","文章內容","文章網址"]
                article_info = []

                for title in titles_list:                
                    y = targets[title]
                    y= ILLEGAL_CHARACTERS_RE.sub(r'', y)
                    article_info.append(y)

                    # print(title,':',y)            
                num+=1
                ws.append(article_info)
                print(web + '已完成寫入' + str(num) + '篇文章')
                break
        except KeyError as k :
            print(k)       
        except AttributeError as attr:
            print(attr)
        except NameError as name:
            print(name)
        except IndexError as index:
            print(index)
        except JSONDecodeError :
            print(JSONDecodeError)

    wb.save(keyword + '_'+ web + '_網路爬蟲.xlsx')
    import pandas as pd
    df = pd.read_excel(keyword + '_'+ web + '_網路爬蟲.xlsx') #請自行依檔案位置調整
    print(keyword+'_'+web+ '_總共有'+str(len(df))+'篇文章')

    #華人健康網爬蟲
    from bs4 import BeautifulSoup
    # from urllib.request#用於獲取網頁
    import requests
    import json
    from openpyxl import Workbook
    import parsel
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
    import re

    ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

    wb = Workbook()
    ws = wb.active

    title =["文章標題",'文章內容','發佈日期']

    ws.append(title)

    web = '華人健康網'
    # 基本參數
    article_url = []
    num_url = 0
    for num in range(1,11):
        try:
            # url = "https://www.top1health.com/Search?q="+keyword+"&page="+str(num)+"&1&type=Illness"
            url = 'https://www.top1health.com/Search?q='+ keyword +'&page='+str(num)+'&type=All'
            rs = requests.session()

            # print(url)
            # get取得頁面的HTML
            response = requests.get(url, headers=headers) 
            response.encoding = 'utf8'
            soup = BeautifulSoup(response.text, "html.parser")
            # print(soup.text) #取得頁面資料
            links = soup.find_all(class_="post-title") #center-block ul li a
            # print(links)
            for link in links:
                if link.a != None:
                    page_url = "https://www.top1health.com/"+link.a["href"]
                    #------這裡有一個cookies防火牆----
                    response_page = requests.get(page_url, headers=headers)
                    result = BeautifulSoup(response_page.text, "html.parser")
                    # print(rponse_pagees)  #確認拿到的網址，等等要進入文章 200
                    # article_data["articleURL"] = page_url
                    article_url.append(page_url)
                    num_url+=1
                    print(web + '已完成寫入' + str(num_url) + '個網址')
        except KeyError as k :
            print(k)        
        except AttributeError as attr:
            print(attr)
        except NameError as name:
            print(name)
        except IndexError as index:
            print(index)
        except JSONDecodeError :
            print(JSONDecodeError)

    #取得所有文章的網址
    # print(article_url)
    print('總共獲取了',len(article_url),'個網址')
    # print(article_url)

    num_article=0
    for url in article_url:
        try:

            # url = "https://www.top1health.com/Search?q="+KEY+"&page="+str(num)+"&1&type=Illness"
            rs = requests.session()

            response = requests.get(url, headers=headers) 
            response.encoding = 'utf8'
            soup = BeautifulSoup(response.text, "html.parser")
            # print(soup.text) #取得頁面資料
            article_info =[]
            
            headline = soup.find("h2") #center-block ul li a
            article_headline = headline.text
            article_info.append(article_headline)
            
            contents = soup.find_all("p") #center-block ul li a
            article_content=''
            for content in contents:
                article_content += content.text
            
            article_content= ILLEGAL_CHARACTERS_RE.sub(r'', article_content)
            article_info.append(article_content)

            selector = parsel.Selector(response.text) # 把獲取下來的html字符串數據，轉成selector可解析的對象
            # print (selector) #<Selector xpath=None data='<html lang="zh-CN" class="ua-mac ua-w...'>
            # css選擇器：就是根據標籤屬性內容，提取相關數據
            # .後面+class名稱 沒有.代表是標籤名稱 :nth-child(i) 表示取這個標籤的第i個
            # 只有一筆資料用get() 標籤內含有多種我們要的資料用getall()
            info = selector.css('.info') #第一次提取，獲取所有li標籤 返回列表
            published_time = info.css('span:nth-child(2)::text').get()
            # print(title)
            article_info.append(published_time)
            # print(article_info)
            ws.append(article_info)
            num_article+=1
            print(web + '已完成寫入' + str(num_article) + '篇文章')
            # break
        except KeyError as k :
            print(k)
        except AttributeError as attr:
            print(attr)
        except NameError as name:
            print(name)
        except IndexError as index:
            print(index)
        except JSONDecodeError :
            print(JSONDecodeError)

    wb.save(keyword + '_yet'+'_'+ web + '_網路爬蟲.xlsx')

    import pandas as pd
    df = pd.read_excel(keyword + '_yet'+'_'+ web +'_網路爬蟲.xlsx') #請自行依檔案位置調整
    # print(keyword+'_'+web+ '_總共有'+str(len(df))+'篇文章')

    cd1 = df.duplicated('文章內容').sum()
    print(web +'重複文章有'+str(cd1))

    # #以文章標題作為標的 刪除重複值
    drop_data = df.drop_duplicates(subset=['文章內容'])

    drop_data.to_excel(keyword+'_ok'+'_'+ web +'_網路爬蟲.xlsx',index=False)
    print(web + '總共文章有'+str(len(drop_data)))



    import requests # 訪問
    import jieba
    from collections import Counter # 次數統計
    import re
    from tkinter import font
    import matplotlib.pyplot as plt
    from wordcloud import WordCloud
    import numpy
    from PIL import Image # 圖片轉array陣列
    import pandas as pd
    import csv
    import random
    from wordcloud import WordCloud, STOPWORDS, ImageColorGenerator  
    from matplotlib import colors
    # titles = ['發佈日期','文章標題','文章內容',"留言內容",'關鍵字',
    #         "文章ID","作者",'看板',"推推留言",'平平留言','噓噓留言',
    #         "文章分類","文章網址",'留言數','喜歡數','標籤']

    web = 'total'
    target_content = '文章標題'
    target_content2 = '文章內容'
    target_content3 = '留言內容'

    all_article = keyword +'_網路爬蟲.txt'
    filename1 = keyword+'_媽咪拜_網路爬蟲.xlsx'
    title =["文章標題","文章內容"]
    df = pd.read_excel(filename1)#,sheetname ='Sheet',header=None)
    df1=df[target_content]
    df2=df[target_content2]
    # df3=df[target_content3]
    df_a=pd.concat([df1,df2])
    a = len(df1)
    # df4.to_csv(a,header=None,sep=',',index=False)

    filename2 = keyword+'_媽咪愛_網路爬蟲.xlsx'
    title = ["文章標題","文章內容","留言內容"]
    df = pd.read_excel(filename2)#,sheetname ='Sheet',header=None)
    df1=df[target_content]
    df2=df[target_content2]
    df3=df[target_content3]
    df_b=pd.concat([df1,df2,df3])
    b = len(df1)
    # df4.to_csv(a,header=None,sep=',',index=False)

    filename3 =keyword+'_ok_Dcard深卡_網路爬蟲.xlsx'
    title = ['文章標題','文章內容']
    df = pd.read_excel(filename3)#,sheetname ='Sheet',header=None)
    df1=df[target_content]
    df2=df[target_content2]
    # df3=df[target_content3]
    df_c=pd.concat([df1,df2])
    c = len(df1)

    # df4.to_csv(a,header=None,sep=',',index=False)

    filename4 =  keyword+'_PTT_網路爬蟲.xlsx' 
    title =["文章標題","文章內容","推推留言",'平平留言','噓噓留言']
    df = pd.read_excel(filename4)#,sheetname ='Sheet',header=None)
    df1=df[target_content]
    df2=df[target_content2]
    df3_1=df["推推留言"]
    df3_2=df["平平留言"]
    df3_3=df["噓噓留言"]
    # df3=df[target_content3]
    df_d=pd.concat([df1,df2,df3_1,df3_2,df3_3])
    d = len(df1)

    # df4.to_csv(a,header=None,sep=',',index=False)

    filename5 = keyword+'_TVBS_網路爬蟲.xlsx'
    title =["文章標題","文章內容"]
    df = pd.read_excel(filename5)#,sheetname ='Sheet',header=None)
    df1=df[target_content]
    df2=df[target_content2]
    # df3=df[target_content3]
    df_e=pd.concat([df1,df2])
    e = len(df1)

    # df4.to_csv(a,header=None,sep=',',index=False)

    filename6 =  keyword +'_ok_華人健康網_網路爬蟲.xlsx'
    title =["文章標題",'文章內容']
    df = pd.read_excel(filename6)#,sheetname ='Sheet',header=None)
    df1=df[target_content]
    df2=df[target_content2]
    # df3=df[target_content3]
    df_f=pd.concat([df1,df2])
    f = len(df1)


    df_all = pd.concat([df_a,df_b,df_c,df_d,df_e,df_f])
    df_all.to_csv(all_article,header=None,sep=',',index=False)

    print('媽咪拜總共有',a,'篇文章')
    print('媽咪愛總共有',b,'篇文章')
    print('Dcard深卡總共有',c,'篇文章')
    print('PTT總共有',d,'篇文章')
    print('TVBS總共有',e,'篇文章')
    print('華人健康網總共有',f,'篇文章')
    print('===================')
    print(keyword + '總共有'+str(a+b+c+d+e+f)+'篇文章')

    print('已完成匯出')

    # with open(all_article, encoding="utf-8", errors='ignore') as f:
    #     text = f.read()
    with open(all_article, encoding="utf-8", errors='ignore') as f:
        text = f.read()

# crawler_all(keyword=keyword)
